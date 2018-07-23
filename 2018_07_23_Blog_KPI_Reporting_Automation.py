import pandas as pd
import numpy as np

import datetime
from datetime import date, timedelta

import sys
sys.path.append('/opt/mnt/publicdrive/Analytics/Gerard/Utils/')
#sys.path.append('/Volumes/ugcompanystorage/Company/public/Analytics/Gerard/Utils/GA/')
from GA.GA_obj import GA

import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Border
import os
from itertools import islice
def find_row(sheet):
    row = 1
    col = 1
    active_cell = 0

    while(active_cell != None):
        row = row + 1
        print(row)
        active_cell = (sheet.cell(row=row, column=col)).value
        print (active_cell)

    print('---')
    print(row)
    return row

def ga_query(start_date, end_date, filter_var, metrics, dimensions, max_results, segment, sort):

    ## The GA object takes  a profile ID and the location of your credential file as argument to create the object
    External = GA('Universal API Key Here', filelocation='/opt/mnt/publicdrive/Analytics/Gerard/Utils/GA/')
    #External = GA('Universal API Key Here', filelocation='/Volumes/ugcompanystorage/Company/public/Analytics/Gerard/Utils/GA/')
    query_response = External.get_results(start_date=start_date,
                    end_date=end_date,
                    filter_var=(None if filter_var==0 else filter_var),
                    metrics=(None if metrics==0 else metrics),
                    dimensions=(None if dimensions==0 else dimensions),
                    max_results=(None if max_results==0 else max_results),
                    segment=(None if segment==0 else segment),
                    sort=(None if sort==0 else sort))

    # Results are stored in 'rows'
    try:

        query_response['rows']

    # coverts list to a dataframe and grabs the first value from the dataframe
        result = pd.DataFrame(query_response['rows']).iloc[0, 0]

    # result:
        return result

    except:
        return 0
#----------------------------------------------------------------------------------------------------------------------
Todays_Date = pd.to_datetime('today')


#Source_Folder_Path = "/Volumes/ugcompanystorage/Company/marketing/social media/blog/Reports/Revenue_Tracking"
Source_Folder_Path = "/opt/mnt/marketingdrive/social media/blog/Reports/Revenue_Tracking"
os.chdir(Source_Folder_Path)


Source_Report_XLSX_File = openpyxl.load_workbook("Source_Report.xlsx")

Source_Report = Source_Report_XLSX_File.get_sheet_by_name('Sheet1')

#
SourceReportDataFrameFull = pd.DataFrame(Source_Report.values)
#print(SourceReportDataFrameFull)

SourceReportDataFrameLessColumnHeaders = SourceReportDataFrameFull.drop(0,axis=0)



#This counts the number of rows in the dataframe
DataFrame_Number_Rows = int(len(SourceReportDataFrameLessColumnHeaders.index))
print(DataFrame_Number_Rows)



#open results file and clear out all data before loop so new up to date data can be stored
Results_File = openpyxl.load_workbook('Blog_Post_Revenue_Tracking_Results.xlsx')

Results_Tab = Results_File.get_sheet_by_name('Results')

for row in Results_Tab['A1:G50000']:
    for cell in row:
        cell.value = None


#column 1 name Date Published
col = 1
Results_Tab.cell(row = 1, column = col).value = str("Date Published")


#column 1 name Post URL
col = col + 1
Results_Tab.cell(row = 1, column = col).value = str("Post URL")

#column 2 name Page_Views_From_Blog
col = col + 1
Results_Tab.cell(row = 1, column = col).value = str("Page_Views_From_Blog")

#column 3 name From_Blog_To_Page_To_Cart"
col = col + 1
Results_Tab.cell(row = 1, column = col).value = str("From_Blog_To_Page_To_Cart")

#column 4 name Item_Revenue
# col = col + 1
# Results_Tab.cell(row = 1, column = col).value = str("Item_Revenue")

#column 5 name Total_Transaction_Revenue
col = col + 1
Results_Tab.cell(row= 1, column=col).value = str("Total_Transaction_Revenue")

Results_File.save('Blog_Post_Revenue_Tracking_Results.xlsx')

#Loop over the source report df, row by row, and assign query paramaters to each variable. Add them to GA queries, and save
#them to the results file

x=0
while x <= DataFrame_Number_Rows:


        query_parameters = SourceReportDataFrameLessColumnHeaders.iloc[x]
        print(query_parameters)

        # = get_query_parameters(DataFrame_Number_Rows)

        #print(all_Parameters)
        
        Start_Date_Raw =  str(query_parameters.iloc[0])
        Filter_1_Raw = query_parameters.iloc[2]
        Regex_Added_To_Cart = query_parameters.iloc[3]




        Start_Date = datetime.datetime.strptime(Start_Date_Raw, '%Y%m%d')
        #Start_Date = Start_Date_Datetime.strftime('%Y-%m-%d')
        print(Start_Date)


        Filter_1 = Filter_1_Raw.split('.com', 1)[1]
        Page_Path_To_Site = "ga:pagePath=@{}".format(Filter_1)
        print(Page_Path_To_Site)

        Combined_filters = "{0};ga:pagePath=~{1}".format(Page_Path_To_Site, Regex_Added_To_Cart)
        print(Combined_filters)

        Revenue_segment = "sessions::condition::{}".format(Page_Path_To_Site)
        print(Revenue_segment)


        #------------------------------------------------------------------------------------------------------
        #metric one Pageviews from blog
        start_date = Start_Date.strftime('%Y-%m-%d')
        end_date = Todays_Date.strftime('%Y-%m-%d')
        filter_var = Page_Path_To_Site
        metrics = 'ga:uniquePageviews'
        dimensions = 0
        max_results = 0
        segment=0
        sort=0

        Page_Views_From_Blog = ga_query(start_date, end_date, filter_var, metrics, dimensions, max_results, segment, sort)
        print('\n1. GA QUERY1 Page_Views_From_Blog: '+ str(Page_Views_From_Blog ))



        #------------------------------------------------------------------------------------------------------

        #metric two added to carts from the blog post
        start_date = Start_Date.strftime('%Y-%m-%d')
        end_date = Todays_Date.strftime('%Y-%m-%d')
        filter_var = Combined_filters
        metrics = 'ga:uniquePageviews'
        dimensions = 0
        max_results = 0
        segment=0
        sort=0

        From_Blog_To_Page_To_Cart = ga_query(start_date, end_date, filter_var, metrics, dimensions, max_results,
                                             segment, sort)
        print('\n2. GA QUERY1 From_Blog_To_Cart: ' + str(From_Blog_To_Page_To_Cart))

        # ------------------------------------------------------------------------------------------------------



        # metric four total transaction revenue from people who viewed the post, navigated to the site, and bought stuff

        #commenting out until can confirm correct ga query for this metric
        # start_date = Start_Date.strftime('%Y-%m-%d')
        # end_date = Todays_Date.strftime('%Y-%m-%d')
        # filter_var = 0
        # metrics = 'ga:itemRevenue'
        # dimensions = 0
        # max_results = 0
        # segment = Revenue_segment
        # sort = 0
        #
        # Item_Revenue = ga_query(start_date, end_date, filter_var, metrics, dimensions, max_results,
        #                                      segment, sort)
        # print('\n3. GA QUERY1 Item_Revenue: ' + str(Item_Revenue))

        # ------------------------------------------------------------------------------------------------------
        # metric three blog post item specific revenue from people who viewed the post and added to cart
        start_date = Start_Date.strftime('%Y-%m-%d')
        end_date = Todays_Date.strftime('%Y-%m-%d')
        filter_var = 0
        metrics = 'ga:transactionRevenue'
        dimensions = 0
        max_results = 0
        segment = Revenue_segment
        sort = 0

        Total_Post_Revenue = ga_query(start_date, end_date, filter_var, metrics, dimensions, max_results,
                                      segment, sort)
        print('\n4. GA QUERY1 Total_Post_Revenue: ' + str(Total_Post_Revenue))
        # ------------------------------------------------------------------------------------------------------


        Results_File = openpyxl.load_workbook('Blog_Post_Revenue_Tracking_Results.xlsx')

        Results_Tab = Results_File.get_sheet_by_name('Results')


        empty_row = find_row(Results_Tab)



        #add item URL
        col = 1
        Results_Tab.cell(row=empty_row, column=col).value = Start_Date.strftime('%Y-%m-%d')

        col = col + 1
        Results_Tab.cell(row = empty_row, column = col).value = Filter_1_Raw

        # add Page_Views_From_Blog to the results file
        col = col + 1
        Results_Tab.cell(row = empty_row, column = col).value = int(Page_Views_From_Blog)
        Results_Tab.cell(row = empty_row, column = col).number_format = "#######"

        # add number of people that came from the blog and added the item to cart
        col = col + 1
        Results_Tab.cell(row = empty_row, column = col).value = int(From_Blog_To_Page_To_Cart)
        Results_Tab.cell(row = empty_row, column = col).number_format = "#######"

        #Add item revenue from people who came from blog and added to cart
        # col = col + 1
        # Results_Tab.cell(row=empty_row, column=col).value = float(Item_Revenue)
        # Results_Tab.cell(row=empty_row, column=col).number_format = "$#######"

        #add total transaction revenue from people who came from the blog and added something on the site to cart
        col = col + 1
        Results_Tab.cell(row=empty_row, column=col).value = float(Total_Post_Revenue)
        Results_Tab.cell(row=empty_row, column=col).number_format = "$#######"



        Results_File.save('Blog_Post_Revenue_Tracking_Results.xlsx')
        # print(test)
        #
        # print(get_query_parameters(DataFrame_Number_Rows))
        # #all_parameters = get_query_parameters(DataFrame_Number_Rows)
        # #print(all_parameters)
        #
        # print(SourceReportDataFrameLessColumnHeaders.iloc[0])



        x=x+1
        #++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
